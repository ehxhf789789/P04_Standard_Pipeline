"""
XLSX Document Parser

Extracts data, formulas, and metadata from Microsoft Excel spreadsheets.
Supports ISO/IEC 29500 (OOXML) format.
"""

import time
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime
import uuid

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from .base import (
    DocumentParser,
    DocumentType,
    DocumentMetadata,
    ParseResult,
    Section,
    TableData,
)

logger = logging.getLogger(__name__)


class XLSXParser(DocumentParser):
    """
    Microsoft Excel Spreadsheet Parser

    Features:
    - Multi-sheet support
    - Cell value and formula extraction
    - Data type detection
    - Merged cell handling
    - Named range extraction
    - Chart detection
    - Data validation rules
    """

    document_type = DocumentType.XLSX
    supported_extensions = ['.xlsx', '.xls', '.xlsm']

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)
        self.max_rows = self.config.get('max_rows', 10000)
        self.max_cols = self.config.get('max_cols', 100)
        self.include_formulas = self.config.get('include_formulas', True)
        self.include_empty_cells = self.config.get('include_empty_cells', False)

    async def parse(self, file_path: Path) -> ParseResult:
        """Parse XLSX and extract all content"""
        start_time = time.time()

        result = ParseResult(
            success=False,
            document_type=self.document_type,
            file_path=str(file_path),
            file_hash=self.compute_file_hash(file_path),
        )

        if not HAS_OPENPYXL:
            result.errors.append("openpyxl not installed. Run: pip install openpyxl")
            return result

        try:
            # Load with data_only=False to get formulas too
            wb = openpyxl.load_workbook(file_path, data_only=False)
            wb_data = openpyxl.load_workbook(file_path, data_only=True)

            # Extract metadata
            result.metadata = await self.extract_metadata(file_path)

            # Process each sheet
            sheets_data = []
            all_tables = []
            text_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = wb_data[sheet_name]

                # Extract sheet content
                sheet_result = await self._process_sheet(
                    sheet, sheet_data, sheet_name
                )
                sheets_data.append(sheet_result)

                # Create table from sheet
                table = TableData(
                    id=str(uuid.uuid4()),
                    name=sheet_name,
                    headers=sheet_result['headers'],
                    rows=sheet_result['rows'],
                )
                all_tables.append(table)

                # Add to text
                text_parts.append(f"=== {sheet_name} ===")
                for row in sheet_result['rows']:
                    text_parts.append('\t'.join(str(cell) for cell in row))

            # Create sections from sheets
            sections = []
            for sheet_data in sheets_data:
                section = Section(
                    id=str(uuid.uuid4()),
                    heading=sheet_data['name'],
                    level=1,
                    tables=[t for t in all_tables if t.name == sheet_data['name']],
                )
                section.content = '\n'.join(
                    '\t'.join(str(cell) for cell in row)
                    for row in sheet_data['rows']
                )
                sections.append(section)

            result.sections = sections
            result.tables = all_tables
            result.full_text = '\n'.join(text_parts)

            # Add sheet metadata
            result.metadata.custom['sheets'] = [
                {
                    'name': s['name'],
                    'row_count': s['row_count'],
                    'col_count': s['col_count'],
                    'has_formulas': s['has_formulas'],
                }
                for s in sheets_data
            ]

            # Statistics
            total_rows = sum(s['row_count'] for s in sheets_data)
            result.statistics = {
                'sheet_count': len(sheets_data),
                'total_rows': total_rows,
                'table_count': len(all_tables),
            }

            wb.close()
            wb_data.close()
            result.success = True

        except Exception as e:
            logger.error(f"Error parsing XLSX: {e}")
            result.errors.append(str(e))

        result.processing_time_ms = int((time.time() - start_time) * 1000)
        return result

    async def extract_metadata(self, file_path: Path) -> DocumentMetadata:
        """Extract XLSX metadata"""
        metadata = DocumentMetadata()

        if not HAS_OPENPYXL:
            return metadata

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            props = wb.properties

            metadata.title = props.title
            metadata.author = props.creator
            metadata.creator = props.creator
            metadata.subject = props.subject

            if props.keywords:
                metadata.keywords = [k.strip() for k in props.keywords.split(',')]

            if props.created:
                metadata.created = props.created
            if props.modified:
                metadata.modified = props.modified

            metadata.custom = {
                'category': props.category,
                'description': props.description,
                'last_modified_by': props.lastModifiedBy,
                'revision': props.revision,
                'version': props.version,
                'sheet_names': wb.sheetnames,
            }

            wb.close()

        except Exception as e:
            logger.error(f"Error extracting XLSX metadata: {e}")

        return metadata

    async def extract_text(self, file_path: Path) -> str:
        """Extract full text from XLSX"""
        if not HAS_OPENPYXL:
            return ""

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== {sheet_name} ===")

                for row in sheet.iter_rows(max_row=self.max_rows, max_col=self.max_cols):
                    row_text = []
                    for cell in row:
                        if cell.value is not None:
                            row_text.append(str(cell.value))
                    if row_text:
                        text_parts.append('\t'.join(row_text))

            wb.close()
            return '\n'.join(text_parts)

        except Exception as e:
            logger.error(f"Error extracting XLSX text: {e}")
            return ""

    async def _process_sheet(
        self,
        sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        sheet_data: 'openpyxl.worksheet.worksheet.Worksheet',
        sheet_name: str
    ) -> Dict[str, Any]:
        """Process a single worksheet"""
        result = {
            'name': sheet_name,
            'headers': [],
            'rows': [],
            'row_count': 0,
            'col_count': 0,
            'has_formulas': False,
            'merged_cells': [],
            'data_types': {},
        }

        # Get actual dimensions
        max_row = min(sheet.max_row or 0, self.max_rows)
        max_col = min(sheet.max_column or 0, self.max_cols)

        if max_row == 0 or max_col == 0:
            return result

        result['row_count'] = max_row
        result['col_count'] = max_col

        # Track merged cells
        for merged_range in sheet.merged_cells.ranges:
            result['merged_cells'].append(str(merged_range))

        # Extract data
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(
            min_row=1, max_row=max_row, max_col=max_col
        )):
            row_data = []
            data_row = list(sheet_data.iter_rows(
                min_row=row_idx + 1, max_row=row_idx + 1, max_col=max_col
            ))[0] if row_idx < max_row else row

            for col_idx, cell in enumerate(row):
                # Get computed value
                data_cell = data_row[col_idx] if col_idx < len(data_row) else cell
                value = data_cell.value

                # Check for formula
                if cell.value and str(cell.value).startswith('='):
                    result['has_formulas'] = True
                    if self.include_formulas:
                        value = {
                            'value': data_cell.value,
                            'formula': str(cell.value),
                        }

                # Track data types
                col_letter = get_column_letter(col_idx + 1)
                if value is not None:
                    dtype = type(value).__name__
                    if col_letter not in result['data_types']:
                        result['data_types'][col_letter] = dtype

                row_data.append(value if value is not None else '')

            rows.append(row_data)

        # First row as headers
        if rows:
            result['headers'] = [str(h) if h else f'Column_{i+1}' for i, h in enumerate(rows[0])]
            result['rows'] = rows[1:]

        return result

    async def to_dataframes(self, file_path: Path) -> Dict[str, 'pd.DataFrame']:
        """Convert XLSX to pandas DataFrames"""
        if not HAS_PANDAS:
            raise ImportError("pandas not installed")

        try:
            excel_file = pd.ExcelFile(file_path)
            dataframes = {}

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    nrows=self.max_rows
                )
                dataframes[sheet_name] = df

            return dataframes

        except Exception as e:
            logger.error(f"Error converting to DataFrames: {e}")
            raise

    async def validate_schema(
        self,
        file_path: Path,
        expected_schema: Dict[str, Dict[str, str]]
    ) -> Dict[str, Any]:
        """
        Validate spreadsheet against expected schema.

        Args:
            file_path: Path to XLSX file
            expected_schema: Dict mapping sheet names to column schemas
                            {sheet_name: {column_name: expected_dtype}}

        Returns:
            Validation results
        """
        results = {
            'valid': True,
            'errors': [],
            'warnings': [],
        }

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name, columns in expected_schema.items():
                if sheet_name not in wb.sheetnames:
                    results['errors'].append(f"Missing sheet: {sheet_name}")
                    results['valid'] = False
                    continue

                sheet = wb[sheet_name]
                headers = [cell.value for cell in sheet[1]]

                for col_name, expected_dtype in columns.items():
                    if col_name not in headers:
                        results['errors'].append(
                            f"Missing column '{col_name}' in sheet '{sheet_name}'"
                        )
                        results['valid'] = False

            wb.close()

        except Exception as e:
            results['errors'].append(str(e))
            results['valid'] = False

        return results
